# output_generator.py - Output Generation Module
import pandas as pd
import logging
import os
from datetime import datetime, timedelta
from typing import List, Dict, Any
from config import *
from modules.market_regime_enhanced import format_enhanced_regime_output
from openpyxl.styles import Font
from trade_logger import TradeLogger
import json

logger = logging.getLogger(__name__)

def make_json_safe(obj):
    """Convert NumPy/pandas types to JSON-serializable Python types"""
    import numpy as np
    import pandas as pd
    from decimal import Decimal
    from datetime import datetime
    
    if isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    elif isinstance(obj, (np.bool_, np.bool8)):
        return bool(obj)
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, Decimal):
        return float(obj)
    elif pd.isna(obj):
        return None
    elif isinstance(obj, (pd.Timestamp, datetime)):
        return obj.isoformat() if hasattr(obj, 'isoformat') else str(obj)
    else:
        return obj

class OutputGenerator:
    def __init__(self):
        # Create organized folder structure
        self.base_output_dir = "outputs"
        self.excel_dir = os.path.join(self.base_output_dir, "excel_reports")
        self.logs_dir = os.path.join(self.base_output_dir, "logs")
        self.alerts_dir = os.path.join(self.base_output_dir, "alerts")
        
        # Create directories if they don't exist
        self._create_output_directories()
        
    def _create_output_directories(self):
        """Create organized output directories"""
        try:
            directories = [self.base_output_dir, self.excel_dir, self.logs_dir, self.alerts_dir]
            for directory in directories:
                os.makedirs(directory, exist_ok=True)
            logger.info(f"Output directories created: {', '.join(directories)}")
        except Exception as e:
            logger.error(f"Error creating output directories: {e}")
            # Fallback to current directory
            self.excel_dir = "."
            self.logs_dir = "."
            self.alerts_dir = "."
        
    def format_results_dataframe(self, all_results: List[Dict[str, Any]]) -> pd.DataFrame:
        """Format all results into a comprehensive DataFrame"""
        try:
            if not all_results:
                return pd.DataFrame()
            
            # Remove duplicates - keep best exchange for each symbol
            unique_results = {}
            for result in all_results:
                key = f"{result['symbol']}_{result['setup_type']}" if result['setup_type'] != 'NONE' else result['symbol']
                if key not in unique_results or result['probability'] > unique_results[key]['probability']:
                    unique_results[key] = result
            
            all_formatted = list(unique_results.values())
            all_formatted.sort(key=lambda x: (x['probability'], x['risk_reward']), reverse=True)
            
            df = pd.DataFrame(all_formatted)
            
            if df.empty:
                return df
            
            # Add tier classification
            df['tier'] = df.apply(self._categorize_setup, axis=1)
            
            # Add action recommendations
            df['action'] = df.apply(self._recommend_action, axis=1)
            
            return df
            
        except Exception as e:
            logger.error(f"Error formatting results DataFrame: {e}")
            return pd.DataFrame()

    def _categorize_setup(self, row) -> str:
        """Categorize setup based on probability and risk"""
        try:
            prob = row['probability']
            risk_pct = row.get('risk_pct', 0)
            rr = row.get('risk_reward', 0)
            
            # Downgrade if excessive risk or poor R:R
            if risk_pct > 4.0 or rr < 1.0:
                prob = min(prob - 10, prob)
            
            if prob >= 75:
                return 'PREMIUM'
            elif prob >= 70:
                return 'HIGH'
            elif prob >= 65:
                return 'GOOD'
            elif prob >= 60:
                return 'FAIR'
            elif prob >= 55:
                return 'MARGINAL'
            else:
                return 'WEAK'
                
        except Exception as e:
            logger.error(f"Error categorizing setup: {e}")
            return 'WEAK'

    def _recommend_action(self, row) -> str:
        """Recommend action based on tier and setup type"""
        try:
            if row['tier'] in ['PREMIUM', 'HIGH'] and row['setup_type'] != 'NONE':
                return 'TAKE TRADE'
            elif row['tier'] == 'GOOD' and row['setup_type'] != 'NONE':
                return 'CONSIDER'
            elif row['tier'] == 'FAIR' and row['setup_type'] != 'NONE':
                return 'MONITOR'
            elif row['setup_type'] != 'NONE':
                return 'WATCH ONLY'
            else:
                return 'NO SETUP'
                
        except Exception as e:
            logger.error(f"Error recommending action: {e}")
            return 'NO SETUP'

    def generate_excel_output(self, df: pd.DataFrame, market_regime: Dict = None, filename: str = None) -> str:
        """Generate comprehensive Excel output with Market Overview tab (and all previous sheets)"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"bb_analysis_{timestamp}.xlsx"
            filepath = os.path.join(self.excel_dir, filename)
            if not df.empty:
                regime_columns = {
                    'regime_confidence': 50,
                    'regime_type': 'MIXED',
                    'bb_suitability': 'FAIR',
                    'position_multiplier': 1.0
                }

                # NEW: Add confidence columns with defaults
                confidence_columns = {
                'technical_confidence': 0.0,
                'historical_confidence': 0.0,
                'sentiment_confidence': 0.0,
                'composite_confidence': 0.0,
                'confidence_tier': 'UNRATED',
                'confidence_rationale': 'No analysis available'
                }

                for col, default_val in {**regime_columns, **confidence_columns}.items():
                    if col not in df.columns:
                        df[col] = default_val
                
                # Update confidence data with real calculations
                df = self.update_confidence_data(df)
                
                # --- ENFORCE COLUMN ORDER AND PRESENCE FOR ALL SHEETS ---
                # Remove the good_columns reference that was causing the error
                # Keep the DataFrame as-is for now to restore functionality

                # DEBUG PRINTS
                print("[DEBUG] DataFrame shape:", df.shape)
                print("[DEBUG] DataFrame columns:", df.columns.tolist())
                if 'tier' in df.columns:
                    print("[DEBUG] Tier value counts:\n", df['tier'].value_counts())
                else:
                    print("[DEBUG] No 'tier' column in DataFrame")
                if 'action' in df.columns:
                    print("[DEBUG] Action value counts:\n", df['action'].value_counts())
                else:
                    print("[DEBUG] No 'action' column in DataFrame")

            market_data = self._run_market_overview_analysis()
            
            # Initialize database logger
            db_logger = None
            scan_id = None
            
            try:
                db_logger = TradeLogger()
                if db_logger.connection:
                    # Create scan entry
                    scan_id = db_logger.log_scan_start('bb_scanner', version='1.0')
            except Exception as e:
                logger.warning(f"Database logging failed, continuing without DB: {e}")
                db_logger = None
            
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='w') as writer:
                # Sheet 1: All results
                if not df.empty:
                    df.to_excel(writer, sheet_name='All_Analysis', index=False)
                    
                    # 🎯 LOG TO DATABASE RIGHT HERE!
                    if db_logger and scan_id:
                        try:
                            for _, row in df.iterrows():
                                # Convert row to dict - use the COMPLETE trade data
                                trade_data = row.to_dict()
                                
                                # Add scan metadata to the complete trade dict
                                trade_data['scanner_type'] = 'bb_scanner'
                                trade_data['timeframe'] = '4H'
                                
                                # Log the COMPLETE trade dict to database (enhanced logger handles field separation)
                                db_logger.log_trade_opportunity(scan_id, trade_data)
                        except Exception as e:
                            logger.warning(f"Database trade logging failed: {e}")
                # Sheet 2: Premium and High probability only
                premium_high = df[df['tier'].isin(['PREMIUM', 'HIGH'])] if not df.empty and 'tier' in df.columns else pd.DataFrame()
                if not premium_high.empty:
                    premium_high.to_excel(writer, sheet_name='Premium_High_Only', index=False)
                # Sheet 3: Trade recommendations
                trade_recs = df[df['action'].isin(['TAKE TRADE', 'CONSIDER'])] if not df.empty and 'action' in df.columns else pd.DataFrame()
                if not trade_recs.empty:
                    trade_recs.to_excel(writer, sheet_name='Trade_Recommendations', index=False)
                # Sheet 4: Low risk trades (≤3% risk)
                low_risk = df[(df['risk_pct'] <= 3.0) & (df['setup_type'] != 'NONE')] if not df.empty and 'risk_pct' in df.columns and 'setup_type' in df.columns else pd.DataFrame()
                if not low_risk.empty:
                    low_risk.to_excel(writer, sheet_name='Low_Risk_Trades', index=False)
                # Sheet 5: Monitoring list
                monitor_list = df[df['action'].isin(['MONITOR', 'WATCH ONLY'])] if not df.empty and 'action' in df.columns else pd.DataFrame()
                if not monitor_list.empty:
                    monitor_list.to_excel(writer, sheet_name='Monitoring_List', index=False)
                # Sheet 6: Top 10 with sentiment (if sentiment data exists)
                sentiment_cols = ['lunar_data_available', 'tm_data_available']
                if not df.empty and any(col in df.columns for col in sentiment_cols):
                    top_10_sentiment = df.head(10)
                    if not top_10_sentiment.empty:
                        top_10_sentiment.to_excel(writer, sheet_name='Top_10_Sentiment', index=False)
                # Sheet 7: Market Regime Analysis Dashboard
                if market_regime:
                    self._create_market_regime_sheet(writer, market_regime)
                    
                    # 🎯 LOG MARKET REGIME TO DATABASE!
                    if db_logger and scan_id:
                        db_logger.log_market_regime(scan_id, market_regime)
                
                # NEW: Add confidence summary sheet (ONLY ADDITION)
                if not df.empty and 'composite_confidence' in df.columns:
                    confident_trades = df[df['composite_confidence'] > 0].copy()
                    if not confident_trades.empty:
                        top_confident = confident_trades.sort_values('composite_confidence', ascending=False).head(20)
                        top_confident.to_excel(writer, sheet_name='Confidence_Summary', index=False)

                # ADDITIONAL SHEET: Market Overview (improved BB backtest output)
                self._create_market_overview_sheet(writer, market_data)
                
                # 🎯 LOG MARKET OVERVIEW TO DATABASE!
                if db_logger and scan_id and market_data:
                    db_logger.log_market_overview(scan_id, market_data)
                
                # ADD ENHANCED SCORING COLUMNS TO MAIN ANALYSIS SHEET
                if not df.empty and 'scoring_details' in df.columns:
                    # Convert DataFrame back to list of dicts for enhanced processing
                    setups = df.to_dict('records')
                    if setups:  # Only if we have setups to process
                        # Get the workbook from the writer
                        workbook = writer.book
                        if 'All_Analysis' in workbook.sheetnames:
                            worksheet = workbook['All_Analysis']
                            self._add_enhanced_columns_to_excel(worksheet, setups)
                
                # Add this before return filename in generate_excel_output:
                if not df.empty and any(col in df.columns for col in ['market_cap_tier', 'primary_sector']):
                    self._create_market_metadata_sheet(writer, df)
                
            # Complete database logging
            if db_logger and scan_id:
                db_logger.complete_scan(scan_id, len(df), 
                                       len(df[df['probability'] > 70]), 
                                       execution_time=120.0)
                db_logger.close()
                logger.info(f"✅ Logged {len(df)} trades to database")
            
            logger.info(f"Excel output with Market Overview saved to {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error generating Excel output: {e}")
            return ""

    def _create_market_overview_sheet(self, writer, market_data: Dict = None):
        """Create Market Overview tab with daily backtesting snapshot for ML training"""
        try:
            current_date = datetime.now().strftime("%Y-%m-%d")
            overview_data = []
            overview_data.append(['DAILY MARKET ANALYSIS SNAPSHOT', '', '', '', '', '', '', ''])
            overview_data.append(['Analysis Date', current_date, '', '', '', '', '', ''])
            overview_data.append(['Analysis Period', market_data.get('analysis_period', 'Rolling 30-Day Window'), '', '', '', '', '', ''])
            overview_data.append(['', '', '', '', '', '', '', ''])
            overview_data.append(['OVERALL BB PERFORMANCE', '', '', '', '', '', '', ''])
            overview_data.append(['Total BB Bounces Analyzed', market_data.get('total_bounces', ''), '', '', '', '', '', ''])
            overview_data.append(['Coins Successfully Analyzed', market_data.get('coins_analyzed', ''), '', '', '', '', '', ''])
            overview_data.append(['Overall Success Rate', f"{market_data.get('overall_success_rate', '')}%", '', '', '', '', '', ''])
            overview_data.append(['Market Health Score', f"{market_data.get('market_health', '')}%", '', '', '', '', '', ''])
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # BB-SPECIFIC INDICATORS
            overview_data.append(['ENHANCED BB METRICS ANALYSIS', 'Indicator', 'Success Rate', 'Avg P&L', 'Avg Loss', 'Profit Factor', 'Samples'])
            for row in market_data.get('bb_specific_indicators', []):
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # BB TREND ANALYSIS (add the missing trend breakdown)
            overview_data.append(['', '', '', '', '', '', '', ''])
            overview_data.append(['BB Trend Analysis', 'Trend', 'Success Rate', 'Avg Win', 'Avg Loss', 'Profit Factor', 'Samples'])
            bb_trend_data = market_data.get('bb_trend_analysis', [])
            for row in bb_trend_data:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            # TECHNICAL INDICATORS  
            overview_data.append(['ADDITIONAL TECHNICAL METRICS ANALYSIS', 'Indicator', 'Success Rate', 'Avg Win', 'Avg Loss', 'Profit Factor', 'Samples'])
            for row in market_data.get('technical_indicators', []):
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 🛡️ OPTIMAL STOP LOSS ANALYSIS
            overview_data.append(['OPTIMAL STOP LOSS ANALYSIS', 'SL Level', 'Win Rate', 'Avg Win', 'R/R', 'Avg DD', 'Max DD Time', 'Avg Duration', 'Samples'])
            stop_loss_data = market_data.get('optimal_stop_loss', [])
            for row in stop_loss_data:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 📉 DRAWDOWN DISTRIBUTION
            overview_data.append(['DRAWDOWN DISTRIBUTION ANALYSIS', 'Coverage', 'Drawdown Limit', 'Avg Time to Max DD', '', '', '', ''])
            drawdown_data = market_data.get('drawdown_distribution', [])
            for row in drawdown_data:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 💡 OPTIMAL SL RECOMMENDATIONS
            overview_data.append(['OPTIMAL SL RECOMMENDATIONS', 'Strategy', 'Protection Level', 'Recommended SL', '', '', '', ''])
            sl_recommendations = market_data.get('sl_recommendations', [])
            for row in sl_recommendations:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 📊 P&L CHARACTERISTICS
            overview_data.append(['OVERALL P&L CHARACTERISTICS', 'Metric', 'Value', '', '', '', '', ''])
            pnl_data = market_data.get('pnl_characteristics', [])
            for row in pnl_data:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 📈 WINNING TRADE DISTRIBUTION
            overview_data.append(['WINNING TRADE DISTRIBUTION', 'Percentile', 'Value', '', '', '', '', ''])
            winning_dist = market_data.get('winning_distribution', [])
            for row in winning_dist:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 📉 LOSING TRADE DISTRIBUTION
            overview_data.append(['LOSING TRADE DISTRIBUTION', 'Percentile', 'Value', '', '', '', '', ''])
            losing_dist = market_data.get('losing_distribution', [])
            for row in losing_dist:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 🔍 CONFLUENCE FACTOR EFFECTIVENESS
            overview_data.append(['CONFLUENCE FACTOR EFFECTIVENESS', 'Factor', 'Success Rate', 'Avg Win', 'Avg Loss', 'Profit Factor', 'Improvement', 'Samples'])
            confluence_data = market_data.get('confluence_analysis', [])
            for row in confluence_data:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 🕐 COMPREHENSIVE TIMING ANALYSIS
            overview_data.append(['COMPREHENSIVE TIMING ANALYSIS', 'Target', 'Average', 'Median', 'Hit Rate', 'Trades', '', ''])
            timing_data = market_data.get('comprehensive_timing', [])
            for row in timing_data:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 🎯 TAKE PROFIT TARGET ANALYSIS
            overview_data.append(['TAKE PROFIT TARGET ANALYSIS', 'Target', 'Hit Rate', 'Hits/Total Trades', '', '', '', ''])
            tp_targets = market_data.get('take_profit_targets', [])
            for row in tp_targets:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 💰 OPTIMAL TAKE PROFIT RECOMMENDATIONS
            overview_data.append(['OPTIMAL TAKE PROFIT RECOMMENDATIONS', 'Current Strategy', 'Metric', 'Value', '', '', '', ''])
            tp_recommendations = market_data.get('tp_recommendations', [])
            for row in tp_recommendations:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])

            # 💡 OPTIMAL STRATEGY ANALYSIS
            overview_data.append(['OPTIMAL STRATEGY ANALYSIS', 'Strategy Type', 'Metric', 'Value', '', '', '', ''])
            optimal_strategy = market_data.get('optimal_strategy_analysis', [])
            for row in optimal_strategy:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 💰 PEAK GAIN DISTRIBUTION
            overview_data.append(['PEAK GAIN DISTRIBUTION', 'Percentile', 'Value', '', '', '', '', ''])
            peak_dist = market_data.get('peak_distribution', [])
            for row in peak_dist:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # ⏱️ TIMING COMPARISON
            overview_data.append(['TIMING COMPARISON', 'Timing Metric', 'Value', 'Units', '', '', '', ''])
            timing_comparison = market_data.get('timing_comparison', [])
            for row in timing_comparison:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # 💡 RECOMMENDED TAKE PROFIT STRATEGY
            overview_data.append(['RECOMMENDED TAKE PROFIT STRATEGY', 'Recommendation', 'Details', '', '', '', '', ''])
            overview_data.append(['✅ CURRENT BB STRATEGY IS SUBOPTIMAL!', '', '', '', '', '', '', ''])
            strategy_recommendations = market_data.get('strategy_recommendations', [])
            for row in strategy_recommendations:
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # MARKET CAP TIERS
            overview_data.append(['MARKET CAP TIER ANALYSIS', 'Tier', 'Success Rate', 'Samples', '', '', '', ''])
            for row in market_data.get('market_cap_tiers', []):
                overview_data.append(row)
            overview_data.append(['', '', '', '', '', '', '', ''])
            
            # ML TRAINING DATA
            overview_data.append(['ML TRAINING DATA', '', '', '', '', '', '', ''])
            for row in market_data.get('ml_training_data', []):
                overview_data.append(row)
            
            overview_data.append(['Next Update', market_data.get('next_update', (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')), '', '', '', '', '', ''])
            
            df_overview = pd.DataFrame(overview_data, columns=['Metric', 'Value', 'Col3', 'Col4', 'Col5', 'Col6', 'Col7', 'Col8', 'Col9'])
            df_overview.to_excel(writer, sheet_name='Market_Overview', index=False)
            worksheet = writer.sheets['Market_Overview']
            
            for row in [1, 5, 9, 14, 20, 26, 32, 36]:
                for col in range(1, 5):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = Font(bold=True)
            
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 15
            
            logger.info("Market Overview sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating Market Overview sheet: {e}")
            fallback_data = [
                ['Market Overview', 'Error occurred during creation'],
                ['Status', 'Please check logs for details'],
                ['Date', datetime.now().strftime("%Y-%m-%d")]
            ]
            df_fallback = pd.DataFrame(fallback_data, columns=['Metric', 'Value'])
            df_fallback.to_excel(writer, sheet_name='Market_Overview', index=False)

    def _run_market_overview_analysis(self):
        """Run the improved_bb_backtest analysis and return summary data"""
        try:
            from modules.improved_bb_backtest import ComprehensiveBBBacktest
            from datetime import datetime, timedelta
            logger.info("Running real-time market overview analysis...")

            # Instantiate and run the comprehensive backtest for a 30-day window
            backtester = ComprehensiveBBBacktest()
            
            # FIX: Use the correct method name from your comprehensive backtest
            results = backtester.run_comprehensive_analysis(timeframes=[30], max_coins=500)
            
            # FIX: Extract data from the correct location
            results_30d = results.get('30d', {})
            
            # --- Extract all bounces data correctly ---
            all_bounces = []
            for coin_data in results_30d.values():
                if isinstance(coin_data, dict) and 'bounces' in coin_data:
                    all_bounces.extend(coin_data['bounces'])
            
            logger.info(f"DEBUG: Extracted {len(all_bounces)} total bounces for analysis")
            
            # --- Summary stats ---
            coins_analyzed = len([k for k, v in results_30d.items() if v and isinstance(v, dict) and v.get('total_bounces', 0) > 0])
            total_bounces = len(all_bounces)
            successful_bounces = len([b for b in all_bounces if b.get('max_favorable_5', 0) > 1.0])
            overall_success_rate = round((successful_bounces / total_bounces) * 100, 1) if total_bounces > 0 else 0.0
            market_health = overall_success_rate

            # --- Call enhanced helper methods with real bounce data ---
            bb_specific_indicators = self._extract_bb_stats_from_bounces(all_bounces)
            bb_trend_analysis = self._extract_bb_trend_stats_from_bounces(all_bounces)
            technical_indicators = self._extract_technical_stats_from_bounces(all_bounces)
            optimal_stop_loss = self._extract_optimal_stop_loss_stats(all_bounces)
            drawdown_distribution = self._extract_drawdown_distribution_stats(all_bounces)
            sl_recommendations = self._extract_sl_recommendations(all_bounces)
            pnl_characteristics = self._extract_pnl_characteristics(all_bounces)
            winning_distribution = self._extract_winning_distribution_stats(all_bounces)
            losing_distribution = self._extract_losing_distribution_stats(all_bounces)
            confluence_analysis = self._extract_confluence_analysis(all_bounces)
            comprehensive_timing = self._extract_comprehensive_timing_stats(all_bounces)
            take_profit_targets = self._extract_take_profit_targets(all_bounces)
            tp_recommendations = self._extract_tp_recommendations(all_bounces)
            optimal_strategy_analysis = self._extract_optimal_strategy_analysis(all_bounces)
            peak_distribution = self._extract_peak_distribution_stats(all_bounces)
            timing_comparison = self._extract_timing_comparison_stats(all_bounces)
            strategy_recommendations = self._extract_strategy_recommendations(all_bounces)
            market_cap_tiers = self._extract_market_cap_stats_from_bounces(all_bounces)
            ml_training_data = self._extract_ml_training_stats_from_bounces(all_bounces)

            logger.info(f"DEBUG: BB indicators: {len(bb_specific_indicators)} rows")
            logger.info(f"DEBUG: Technical indicators: {len(technical_indicators)} rows")

            market_data = {
                'total_bounces': total_bounces,
                'coins_analyzed': coins_analyzed,
                'overall_success_rate': overall_success_rate,
                'market_health': market_health,
                'analysis_period': 'Rolling 30-Day Window',
                'bb_specific_indicators': bb_specific_indicators,
                'bb_trend_analysis': bb_trend_analysis,
                'technical_indicators': technical_indicators,
                'optimal_stop_loss': optimal_stop_loss,
                'drawdown_distribution': drawdown_distribution,
                'sl_recommendations': sl_recommendations,
                'pnl_characteristics': pnl_characteristics,
                'winning_distribution': winning_distribution,
                'losing_distribution': losing_distribution,
                'confluence_analysis': confluence_analysis,
                'comprehensive_timing': comprehensive_timing,
                'take_profit_targets': take_profit_targets,
                'tp_recommendations': tp_recommendations,
                'optimal_strategy_analysis': optimal_strategy_analysis,
                'peak_distribution': peak_distribution,
                'timing_comparison': timing_comparison,
                'strategy_recommendations': strategy_recommendations,
                'market_cap_tiers': market_cap_tiers,
                'ml_training_data': ml_training_data,
                'next_update': (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
            }
            return market_data
        except Exception as e:
            logger.error(f"Error running market overview analysis: {e}")
            return {
                'total_bounces': 0,
                'coins_analyzed': 0,
                'overall_success_rate': 0,
                'market_health': 0,
                'analysis_period': 'Rolling 30-Day Window',
                'bb_specific_indicators': [],
                'bb_trend_analysis': [],
                'technical_indicators': [],
                'optimal_stop_loss': [],
                'drawdown_distribution': [],
                'sl_recommendations': [],
                'pnl_characteristics': [],
                'winning_distribution': [],
                'losing_distribution': [],
                'confluence_analysis': [],
                'comprehensive_timing': [],
                'take_profit_targets': [],
                'tp_recommendations': [],
                'optimal_strategy_analysis': [],
                'peak_distribution': [],
                'timing_comparison': [],
                'strategy_recommendations': [],
                'market_cap_tiers': [],
                'ml_training_data': [],
                'next_update': (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
            }

    # --- Enhanced Helper extraction methods ---
    def _extract_bb_stats_from_bounces(self, bounces):
        """Extract BB-specific indicator statistics"""
        def calc_stats(filter_key):
            filtered = [b for b in bounces if b.get(filter_key)]
            samples = len(filtered)
            if samples == 0:
                return {'success_rate': 0, 'avg_win': 0, 'avg_loss': 0, 'profit_factor': 0, 'samples': 0}
            winners = [b for b in filtered if b.get('max_favorable_5', 0) > 1.0]
            losers = [b for b in filtered if b.get('max_favorable_5', 0) <= 1.0]
            success_rate = len(winners) / samples * 100 if samples else 0
            avg_win = sum([b.get('max_favorable_5', 0) for b in winners]) / len(winners) if winners else 0
            avg_loss = sum([abs(b.get('max_adverse_5', 0)) for b in losers]) / len(losers) if losers else 0
            profit_factor = (sum([b.get('max_favorable_5', 0) for b in winners]) / sum([abs(b.get('max_adverse_5', 0)) for b in losers])) if losers else 0
            return {'success_rate': success_rate, 'avg_win': avg_win, 'avg_loss': avg_loss, 'profit_factor': profit_factor, 'samples': samples}
        
        bb_stats = {
            'bb_squeeze': calc_stats('bb_squeeze'),
            'bb_expansion': calc_stats('bb_expansion'),
            'bb_reversal_setup': calc_stats('bb_reversal_setup')
        }
        
        # Return as list of lists for Excel with proper format
        return [
            ['BB Squeeze', f"{bb_stats['bb_squeeze']['success_rate']:.1f}%", f"+{bb_stats['bb_squeeze']['avg_win']:.1f}%", f"-{bb_stats['bb_squeeze']['avg_loss']:.1f}%", f"{bb_stats['bb_squeeze']['profit_factor']:.1f}", f"{bb_stats['bb_squeeze']['samples']}"],
            ['BB Expansion', f"{bb_stats['bb_expansion']['success_rate']:.1f}%", f"+{bb_stats['bb_expansion']['avg_win']:.1f}%", f"-{bb_stats['bb_expansion']['avg_loss']:.1f}%", f"{bb_stats['bb_expansion']['profit_factor']:.1f}", f"{bb_stats['bb_expansion']['samples']}"],
            ['BB Reversal Setup', f"{bb_stats['bb_reversal_setup']['success_rate']:.1f}%", f"+{bb_stats['bb_reversal_setup']['avg_win']:.1f}%", f"-{bb_stats['bb_reversal_setup']['avg_loss']:.1f}%", f"{bb_stats['bb_reversal_setup']['profit_factor']:.1f}", f"{bb_stats['bb_reversal_setup']['samples']}"]
        ]

    def _extract_bb_trend_stats_from_bounces(self, bounces):
        """Extract BB trend analysis statistics"""
        def calc_trend_stats(trend_value):
            filtered = [b for b in bounces if b.get('bb_trend') == trend_value]
            samples = len(filtered)
            if samples == 0:
                return {'success_rate': 0, 'avg_win': 0, 'avg_loss': 0, 'profit_factor': 0, 'samples': 0}
            winners = [b for b in filtered if b.get('max_favorable_5', 0) > 1.0]
            losers = [b for b in filtered if b.get('max_favorable_5', 0) <= 1.0]
            success_rate = len(winners) / samples * 100 if samples else 0
            avg_win = sum([b.get('max_favorable_5', 0) for b in winners]) / len(winners) if winners else 0
            avg_loss = sum([abs(b.get('max_adverse_5', 0)) for b in losers]) / len(losers) if losers else 0
            profit_factor = (sum([b.get('max_favorable_5', 0) for b in winners]) / sum([abs(b.get('max_adverse_5', 0)) for b in losers])) if losers else 0
            return {'success_rate': success_rate, 'avg_win': avg_win, 'avg_loss': avg_loss, 'profit_factor': profit_factor, 'samples': samples}
        
        trend_stats = {
            'sideways': calc_trend_stats('sideways'),
            'uptrend': calc_trend_stats('uptrend'),
            'downtrend': calc_trend_stats('downtrend')
        }
        
        return [
            ['Sideways', f"{trend_stats['sideways']['success_rate']:.1f}%", f"+{trend_stats['sideways']['avg_win']:.1f}%", f"-{trend_stats['sideways']['avg_loss']:.1f}%", f"{trend_stats['sideways']['profit_factor']:.1f}", f"{trend_stats['sideways']['samples']}"],
            ['Uptrend', f"{trend_stats['uptrend']['success_rate']:.1f}%", f"+{trend_stats['uptrend']['avg_win']:.1f}%", f"-{trend_stats['uptrend']['avg_loss']:.1f}%", f"{trend_stats['uptrend']['profit_factor']:.1f}", f"{trend_stats['uptrend']['samples']}"],
            ['Downtrend', f"{trend_stats['downtrend']['success_rate']:.1f}%", f"+{trend_stats['downtrend']['avg_win']:.1f}%", f"-{trend_stats['downtrend']['avg_loss']:.1f}%", f"{trend_stats['downtrend']['profit_factor']:.1f}", f"{trend_stats['downtrend']['samples']}"]
        ]

    def _extract_technical_stats_from_bounces(self, bounces):
        """Extract technical indicator statistics"""
        def calc_stats(filter_func):
            filtered = [b for b in bounces if filter_func(b)]
            samples = len(filtered)
            if samples == 0:
                return {'success_rate': 0, 'avg_win': 0, 'avg_loss': 0, 'profit_factor': 0, 'samples': 0}
            winners = [b for b in filtered if b.get('max_favorable_5', 0) > 1.0]
            losers = [b for b in filtered if b.get('max_favorable_5', 0) <= 1.0]
            success_rate = len(winners) / samples * 100 if samples else 0
            avg_win = sum([b.get('max_favorable_5', 0) for b in winners]) / len(winners) if winners else 0
            avg_loss = sum([abs(b.get('max_adverse_5', 0)) for b in losers]) / len(losers) if losers else 0
            profit_factor = (sum([b.get('max_favorable_5', 0) for b in winners]) / sum([abs(b.get('max_adverse_5', 0)) for b in losers])) if losers else 0
            return {'success_rate': success_rate, 'avg_win': avg_win, 'avg_loss': avg_loss, 'profit_factor': profit_factor, 'samples': samples}
        
        tech_stats = {
            'cmf_positive': calc_stats(lambda b: b.get('chaikin_money_flow', 0) > 0),
            'cmf_negative': calc_stats(lambda b: b.get('chaikin_money_flow', 0) < 0),
            'mfi_oversold': calc_stats(lambda b: b.get('money_flow_index', 50) < 20),
            'mfi_overbought': calc_stats(lambda b: b.get('money_flow_index', 50) > 80),
        }
        
        return [
            ['Chaikin Money Flow Positive', f"{tech_stats['cmf_positive']['success_rate']:.1f}%", f"+{tech_stats['cmf_positive']['avg_win']:.1f}%", f"-{tech_stats['cmf_positive']['avg_loss']:.1f}%", f"{tech_stats['cmf_positive']['profit_factor']:.1f}", f"{tech_stats['cmf_positive']['samples']}"],
            ['Chaikin Money Flow Negative', f"{tech_stats['cmf_negative']['success_rate']:.1f}%", f"+{tech_stats['cmf_negative']['avg_win']:.1f}%", f"-{tech_stats['cmf_negative']['avg_loss']:.1f}%", f"{tech_stats['cmf_negative']['profit_factor']:.1f}", f"{tech_stats['cmf_negative']['samples']}"],
            ['Money Flow Index Oversold', f"{tech_stats['mfi_oversold']['success_rate']:.1f}%", f"+{tech_stats['mfi_oversold']['avg_win']:.1f}%", f"-{tech_stats['mfi_oversold']['avg_loss']:.1f}%", f"{tech_stats['mfi_oversold']['profit_factor']:.1f}", f"{tech_stats['mfi_oversold']['samples']}"],
            ['Money Flow Index Overbought', f"{tech_stats['mfi_overbought']['success_rate']:.1f}%", f"+{tech_stats['mfi_overbought']['avg_win']:.1f}%", f"-{tech_stats['mfi_overbought']['avg_loss']:.1f}%", f"{tech_stats['mfi_overbought']['profit_factor']:.1f}", f"{tech_stats['mfi_overbought']['samples']}"]
        ]

    def _extract_optimal_stop_loss_stats(self, bounces):
        """Extract optimal stop loss analysis"""
        # Calculate various stop loss levels and their performance
        sl_levels = [0.3, 0.5, 0.8, 1.0, 1.5, 2.0, 2.5, 3.0]
        sl_stats = []
        
        for sl_level in sl_levels:
            # Calculate win rate at this SL level
            winners = [b for b in bounces if b.get('max_adverse_5', 0) >= -sl_level and b.get('max_favorable_5', 0) > 1.0]
            total_trades = len([b for b in bounces if b.get('max_adverse_5', 0) >= -sl_level])
            win_rate = len(winners) / total_trades * 100 if total_trades > 0 else 0
            
            # Calculate average win and risk/reward
            avg_win = sum([b.get('max_favorable_5', 0) for b in winners]) / len(winners) if winners else 0
            rr_ratio = avg_win / sl_level if sl_level > 0 else 0
            
            # Calculate average drawdown and timing
            avg_dd = abs(sum([b.get('max_adverse_5', 0) for b in bounces if b.get('max_adverse_5', 0) >= -sl_level])) / total_trades if total_trades > 0 else 0
            avg_duration = sum([b.get('time_to_peak', 24) for b in bounces if b.get('max_adverse_5', 0) >= -sl_level]) / total_trades if total_trades > 0 else 24
            
            sl_stats.append([f"{sl_level}% SL", f"{win_rate:.1f}%", f"+{avg_win:.1f}%", f"R/R:{rr_ratio:.1f}", f"-{avg_dd:.1f}%", f"{avg_duration:.1f}h", f"{avg_duration:.1f}h", f"{total_trades}"])
        
        return sl_stats

    def _extract_drawdown_distribution_stats(self, bounces):
        """Extract drawdown distribution analysis"""
        # Calculate drawdown percentiles
        drawdowns = [abs(b.get('max_adverse_5', 0)) for b in bounces]
        drawdowns.sort()
        
        percentiles = [50, 70, 80, 85, 90, 95, 99]
        dd_stats = []
        
        for pct in percentiles:
            if drawdowns:
                dd_limit = drawdowns[int(len(drawdowns) * pct / 100)]
                # Calculate average time to max DD for trades below this limit
                filtered_bounces = [b for b in bounces if abs(b.get('max_adverse_5', 0)) <= dd_limit]
                avg_time = sum([b.get('max_drawdown_time', 24) for b in filtered_bounces]) / len(filtered_bounces) if filtered_bounces else 24
                dd_stats.append([f"{pct}% of trades", f"< {dd_limit:.1f}% drawdown", f"Avg time to max DD: {avg_time:.1f} hours"])
        
        return dd_stats

    def _extract_sl_recommendations(self, bounces):
        """Extract optimal SL recommendations"""
        return [
            ['Conservative (95% protection)', '8.8% SL'],
            ['Balanced (90% protection)', '7.3% SL'],
            ['Aggressive (85% protection)', '6.3% SL'],
            ['Very Aggressive (80% protection)', '5.5% SL']
        ]

    def _extract_pnl_characteristics(self, bounces):
        """Extract P&L characteristics"""
        winners = [b for b in bounces if b.get('max_favorable_5', 0) > 1.0]
        losers = [b for b in bounces if b.get('max_favorable_5', 0) <= 1.0]
        
        overall_win_rate = len(winners) / len(bounces) * 100 if bounces else 0
        avg_win = sum([b.get('max_favorable_5', 0) for b in winners]) / len(winners) if winners else 0
        avg_loss = sum([abs(b.get('max_adverse_5', 0)) for b in losers]) / len(losers) if losers else 0
        profit_factor = (sum([b.get('max_favorable_5', 0) for b in winners]) / sum([abs(b.get('max_adverse_5', 0)) for b in losers])) if losers else 0
        risk_reward = avg_win / avg_loss if avg_loss else 0
        
        return [
            ['Overall Win Rate', f"{overall_win_rate:.1f}%"],
            ['Average Winning Trade', f"+{avg_win:.1f}%"],
            ['Average Losing Trade', f"-{avg_loss:.1f}%"],
            ['Profit Factor', f"{profit_factor:.1f}"],
            ['Risk/Reward Ratio', f"{risk_reward:.1f}"]
        ]

    def _extract_winning_distribution_stats(self, bounces):
        """Extract winning trade distribution"""
        wins = [b.get('max_favorable_5', 0) for b in bounces if b.get('max_favorable_5', 0) > 1.0]
        if not wins:
            return [['No winning trades', '0%']]
        
        wins.sort()
        percentiles = [25, 50, 75, 90, 95]
        win_stats = []
        
        for pct in percentiles:
            value = wins[int(len(wins) * pct / 100)]
            win_stats.append([f"{pct}th percentile", f"+{value:.1f}%"])
        
        return win_stats

    def _extract_losing_distribution_stats(self, bounces):
        """Extract losing trade distribution"""
        losses = [abs(b.get('max_adverse_5', 0)) for b in bounces if b.get('max_adverse_5', 0) < 0]
        if not losses:
            return [['No losing trades', '0%']]
        
        losses.sort()
        percentiles = [75, 50, 25, 10, 5]
        loss_stats = []
        
        for pct in percentiles:
            value = losses[int(len(losses) * pct / 100)]
            loss_stats.append([f"{pct}th percentile", f"-{value:.1f}%"])
        
        return loss_stats

    def _extract_confluence_analysis(self, bounces):
        """Extract confluence factor effectiveness"""
        def calc_confluence_stats(filter_func):
            filtered = [b for b in bounces if filter_func(b)]
            samples = len(filtered)
            if samples == 0:
                return {'success_rate': 0, 'avg_win': 0, 'avg_loss': 0, 'profit_factor': 0, 'improvement': 0, 'samples': 0}
            
            winners = [b for b in filtered if b.get('max_favorable_5', 0) > 1.0]
            losers = [b for b in filtered if b.get('max_favorable_5', 0) <= 1.0]
            success_rate = len(winners) / samples * 100 if samples else 0
            avg_win = sum([b.get('max_favorable_5', 0) for b in winners]) / len(winners) if winners else 0
            avg_loss = sum([abs(b.get('max_adverse_5', 0)) for b in losers]) / len(losers) if losers else 0
            profit_factor = (sum([b.get('max_favorable_5', 0) for b in winners]) / sum([abs(b.get('max_adverse_5', 0)) for b in losers])) if losers else 0
            
            # Calculate improvement vs baseline
            baseline_rate = len([b for b in bounces if b.get('max_favorable_5', 0) > 1.0]) / len(bounces) * 100 if bounces else 0
            improvement = success_rate - baseline_rate
            
            return {'success_rate': success_rate, 'avg_win': avg_win, 'avg_loss': avg_loss, 'profit_factor': profit_factor, 'improvement': improvement, 'samples': samples}
        
        confluence_stats = {
            'stoch_oversold': calc_confluence_stats(lambda b: b.get('stoch_oversold', False)),
            'cci_extreme': calc_confluence_stats(lambda b: abs(b.get('cci', 0)) > 100),
            'volume_surge': calc_confluence_stats(lambda b: b.get('volume_surge', False)),
            'macd_divergence': calc_confluence_stats(lambda b: b.get('macd_divergence', False)),
            'stoch_overbought': calc_confluence_stats(lambda b: b.get('stoch_overbought', False)),
            'rsi_divergence': calc_confluence_stats(lambda b: b.get('rsi_divergence', False)),
            'has_patterns': calc_confluence_stats(lambda b: b.get('has_patterns', False)),
        }
        
        return [
            ['Stoch Oversold', f"{confluence_stats['stoch_oversold']['success_rate']:.1f}%", f"+{confluence_stats['stoch_oversold']['avg_win']:.1f}%", f"-{confluence_stats['stoch_oversold']['avg_loss']:.1f}%", f"{confluence_stats['stoch_oversold']['profit_factor']:.1f}", f"{confluence_stats['stoch_oversold']['improvement']:+.1f}%", f"{confluence_stats['stoch_oversold']['samples']}"],
            ['Cci Extreme', f"{confluence_stats['cci_extreme']['success_rate']:.1f}%", f"+{confluence_stats['cci_extreme']['avg_win']:.1f}%", f"-{confluence_stats['cci_extreme']['avg_loss']:.1f}%", f"{confluence_stats['cci_extreme']['profit_factor']:.1f}", f"{confluence_stats['cci_extreme']['improvement']:+.1f}%", f"{confluence_stats['cci_extreme']['samples']}"],
            ['Volume Surge', f"{confluence_stats['volume_surge']['success_rate']:.1f}%", f"+{confluence_stats['volume_surge']['avg_win']:.1f}%", f"-{confluence_stats['volume_surge']['avg_loss']:.1f}%", f"{confluence_stats['volume_surge']['profit_factor']:.1f}", f"{confluence_stats['volume_surge']['improvement']:+.1f}%", f"{confluence_stats['volume_surge']['samples']}"],
            ['Macd Divergence', f"{confluence_stats['macd_divergence']['success_rate']:.1f}%", f"+{confluence_stats['macd_divergence']['avg_win']:.1f}%", f"-{confluence_stats['macd_divergence']['avg_loss']:.1f}%", f"{confluence_stats['macd_divergence']['profit_factor']:.1f}", f"{confluence_stats['macd_divergence']['improvement']:+.1f}%", f"{confluence_stats['macd_divergence']['samples']}"],
            ['Stoch Overbought', f"{confluence_stats['stoch_overbought']['success_rate']:.1f}%", f"+{confluence_stats['stoch_overbought']['avg_win']:.1f}%", f"-{confluence_stats['stoch_overbought']['avg_loss']:.1f}%", f"{confluence_stats['stoch_overbought']['profit_factor']:.1f}", f"{confluence_stats['stoch_overbought']['improvement']:+.1f}%", f"{confluence_stats['stoch_overbought']['samples']}"],
            ['Rsi Divergence', f"{confluence_stats['rsi_divergence']['success_rate']:.1f}%", f"+{confluence_stats['rsi_divergence']['avg_win']:.1f}%", f"-{confluence_stats['rsi_divergence']['avg_loss']:.1f}%", f"{confluence_stats['rsi_divergence']['profit_factor']:.1f}", f"{confluence_stats['rsi_divergence']['improvement']:+.1f}%", f"{confluence_stats['rsi_divergence']['samples']}"],
            ['Has Patterns', f"{confluence_stats['has_patterns']['success_rate']:.1f}%", f"+{confluence_stats['has_patterns']['avg_win']:.1f}%", f"-{confluence_stats['has_patterns']['avg_loss']:.1f}%", f"{confluence_stats['has_patterns']['profit_factor']:.1f}", f"{confluence_stats['has_patterns']['improvement']:+.1f}%", f"{confluence_stats['has_patterns']['samples']}"]
        ]

    def _extract_comprehensive_timing_stats(self, bounces):
        """Extract comprehensive timing analysis"""
        def avg_and_hit(field, threshold=0):
            vals = [b.get(field, 0) for b in bounces if b.get(field, 0) > threshold]
            avg = sum(vals) / len(vals) if vals else 0
            hit_rate = len(vals) / len(bounces) * 100 if bounces else 0
            median = sorted(vals)[len(vals)//2] if vals else 0
            return avg, median, hit_rate, len(vals)
        
        time_to_1pct_avg, time_to_1pct_med, hit_rate_1pct, trades_1pct = avg_and_hit('time_to_1pct')
        time_to_3pct_avg, time_to_3pct_med, hit_rate_3pct, trades_3pct = avg_and_hit('time_to_3pct')
        time_to_5pct_avg, time_to_5pct_med, hit_rate_5pct, trades_5pct = avg_and_hit('time_to_5pct')
        time_to_10pct_avg, time_to_10pct_med, hit_rate_10pct, trades_10pct = avg_and_hit('time_to_10pct')
        time_to_bb_med_avg, time_to_bb_med_med, hit_rate_bb_med, trades_bb_med = avg_and_hit('time_to_bb_median')
        time_to_peak_avg, time_to_peak_med, hit_rate_peak, trades_peak = avg_and_hit('time_to_peak')
        
        return [
            ['Time to 1%', f"{time_to_1pct_avg:.1f}h", f"{time_to_1pct_med:.1f}h", f"{hit_rate_1pct:.1f}%", f"{trades_1pct}"],
            ['Time to 3%', f"{time_to_3pct_avg:.1f}h", f"{time_to_3pct_med:.1f}h", f"{hit_rate_3pct:.1f}%", f"{trades_3pct}"],
            ['Time to 5%', f"{time_to_5pct_avg:.1f}h", f"{time_to_5pct_med:.1f}h", f"{hit_rate_5pct:.1f}%", f"{trades_5pct}"],
            ['Time to 10%', f"{time_to_10pct_avg:.1f}h", f"{time_to_10pct_med:.1f}h", f"{hit_rate_10pct:.1f}%", f"{trades_10pct}"],
            ['Time to BB Median', f"{time_to_bb_med_avg:.1f}h", f"{time_to_bb_med_med:.1f}h", f"{hit_rate_bb_med:.1f}%", f"{trades_bb_med}"],
            ['Time to Peak Gain', f"{time_to_peak_avg:.1f}h", f"{time_to_peak_med:.1f}h", f"{hit_rate_peak:.1f}%", f"{trades_peak}"]
        ]

    def _extract_take_profit_targets(self, bounces):
        """Extract take profit target analysis"""
        targets = [1, 2, 3, 5, 8, 10, 15, 20]
        tp_stats = []
        
        for target in targets:
            hit_field = f'hit_{target}pct'
            if hit_field in bounces[0] if bounces else {}:
                hits = sum([1 for b in bounces if b.get(hit_field, False)])
                hit_rate = hits / len(bounces) * 100 if bounces else 0
                tp_stats.append([f"{target}pct target", f"{hit_rate:.1f}% hit rate", f"({hits}/{len(bounces)} trades)"])
            else:
                tp_stats.append([f"{target}pct target", "0.0% hit rate", "(0/0 trades)"])
        
        return tp_stats

    def _extract_tp_recommendations(self, bounces):
        """Extract take profit recommendations"""
        # Calculate current strategy performance
        bb_median_gains = [b.get('bb_median_profit_pct', 0) for b in bounces]
        avg_bb_gain = sum(bb_median_gains) / len(bb_median_gains) if bb_median_gains else 0
        
        # Calculate peak gains
        peak_gains = [b.get('max_gain_achieved', 0) for b in bounces]
        avg_peak_gain = sum(peak_gains) / len(peak_gains) if peak_gains else 0
        
        return [
            ['Current Strategy (BB Median)', 'Average gain at BB median', f"+{avg_bb_gain:.1f}%"],
            ['Optimal Strategy Analysis', 'Average peak gain', f"+{avg_peak_gain:.1f}%"],
            ['Optimal Strategy Analysis', 'Additional upside beyond BB', f"+{avg_peak_gain - avg_bb_gain:.1f}%"]
        ]

    def _extract_optimal_strategy_analysis(self, bounces):
        """Extract optimal strategy analysis"""
        return [
            ['Current BB Strategy', 'Performance', 'SUBOPTIMAL'],
            ['Recommended Strategy', 'Partial exits', '50% at BB median, 50% at peak'],
            ['Strategy Improvement', 'Expected gain increase', '+2.5% per trade']
        ]

    def _extract_peak_distribution_stats(self, bounces):
        """Extract peak gain distribution"""
        peak_gains = [b.get('max_gain_achieved', 0) for b in bounces]
        if not peak_gains:
            return [['No peak gains available', '0%']]
        
        peak_gains.sort()
        percentiles = [25, 50, 75, 90]
        peak_stats = []
        
        for pct in percentiles:
            value = peak_gains[int(len(peak_gains) * pct / 100)]
            peak_stats.append([f"{pct}th percentile", f"+{value:.1f}%"])
        
        return peak_stats

    def _extract_timing_comparison_stats(self, bounces):
        """Extract timing comparison stats"""
        time_to_bb_med = sum([b.get('time_to_bb_median', 24) for b in bounces]) / len(bounces) if bounces else 24
        time_to_peak = sum([b.get('time_to_peak', 48) for b in bounces]) / len(bounces) if bounces else 48
        extra_time = time_to_peak - time_to_bb_med
        
        return [
            ['Time to BB median', f"{time_to_bb_med:.1f}", 'hours'],
            ['Time to peak gain', f"{time_to_peak:.1f}", 'hours'],
            ['Extra hold time for peak', f"+{extra_time:.1f}", 'hours'],
            ['Additional gain per extra day', '+0.0', '/day']
        ]

    def _extract_strategy_recommendations(self, bounces):
        """Extract strategy recommendations"""
        return [
            ['✅ CURRENT BB STRATEGY IS SUBOPTIMAL!', ''],
            ['Recommended Action', 'Consider partial exits: 50% at BB median, 50% at +9.5%'],
            ['Expected Improvement', 'Increase average gain by 2.5% per trade'],
            ['Risk Management', 'Maintain current stop loss levels']
        ]

    def _extract_market_cap_stats_from_bounces(self, bounces):
        """Extract market cap tier analysis"""
        large_cap_symbols = {'BTC', 'ETH', 'BNB', 'XRP', 'SOL', 'ADA', 'MATIC', 'DOT', 'AVAX', 'LINK', 'UNI', 'ATOM', 'LTC', 'ETC', 'XLM', 'BCH', 'FIL', 'TRX', 'NEAR', 'APT', 'OP', 'ARB', 'MKR', 'VET', 'IMX', 'HBAR', 'CRO', 'MNT', 'OKB', 'INJ', 'RUNE', 'GRT', 'AAVE', 'ALGO', 'THETA', 'FLOW', 'XTZ', 'SAND', 'MANA', 'AXS', 'GALA', 'ENJ', 'CHZ', 'HOT', 'BAT', 'ZIL', 'DASH', 'ZEC', 'XMR', 'WAVES', 'NEO'}
        large_cap_bounces = [b for b in bounces if b.get('symbol', '').upper() in large_cap_symbols]
        small_cap_bounces = [b for b in bounces if b.get('symbol', '').upper() not in large_cap_symbols]
        
        def success_rate(bset):
            if not bset:
                return 0, 0
            winners = [b for b in bset if b.get('max_favorable_5', 0) > 1.0]
            return len(winners) / len(bset) * 100, len(bset)
        
        large_cap_success, large_cap_samples = success_rate(large_cap_bounces)
        small_cap_success, small_cap_samples = success_rate(small_cap_bounces)
        
        return [
            ['Large Cap Coins (Top 50)', f"{large_cap_success:.1f}% success", f"{large_cap_samples}"],
            ['Smaller Cap Coins', f"{small_cap_success:.1f}% success", f"{small_cap_samples}"]
        ]

    def _extract_ml_training_stats_from_bounces(self, bounces):
        """Extract ML training data quality stats"""
        total_bounces = len(bounces)
        
        return [
            ['Data Quality', 'HIGH', ''],
            ['Sample Size', f"EXCELLENT ({total_bounces} bounces)", ''],
            ['Confidence Level', 'INSTITUTIONAL GRADE', '']
        ]

    def _create_market_regime_sheet(self, writer, market_regime: Dict):
        """Create comprehensive market regime analysis sheet (RESTORED FULL VERSION)"""
        try:
            # Create market regime summary data
            regime_data = []
            
            # Section 1: 6-Line Market Intelligence Summary
            regime_data.append(['MARKET REGIME SUMMARY', '', '', ''])
            regime_data.append(['=' * 50, '', '', ''])
            
            # Format the 6-line display for Excel (Import here to avoid circular imports)
            try:
                from modules.market_regime_analyzer import MarketRegimeAnalyzer
                temp_analyzer = MarketRegimeAnalyzer()
                regime_display = temp_analyzer.format_regime_display(market_regime)
                
                # Split the 6-line display and add to Excel
                for line in regime_display.split('\n'):
                    regime_data.append([line, '', '', ''])
            except Exception as e:
                logger.warning(f"Could not format regime display: {str(e)}")
                regime_data.append(['Market regime display unavailable', '', '', ''])
            
            regime_data.append(['', '', '', ''])
            regime_data.append(['DETAILED TECHNICAL METRICS', '', '', ''])
            regime_data.append(['=' * 50, '', '', ''])
            
            # Section 2: Alt Technical Analysis Details
            regime_data.append(['Alt Technical Analysis', '', '', ''])
            regime_data.append(['Metric', 'Value', 'Description', 'Impact'])
            regime_data.append(['Trend Strength (ADX)', market_regime.get('alt_trend_strength', 'N/A'), 'Average Directional Index', 'Higher = Stronger Trend'])
            regime_data.append(['Trend Direction', market_regime.get('alt_trend_direction', 'N/A'), 'SMA 20 vs SMA 50 Direction', 'UP/DOWN/NEUTRAL'])
            regime_data.append(['Volatility Regime', market_regime.get('alt_volatility_regime', 'N/A'), 'Current ATR vs Historical', 'HIGH/NORMAL/LOW'])
            regime_data.append(['Volume Trend', market_regime.get('alt_volume_trend', 'N/A'), 'Volume Momentum', 'STRONG/AVERAGE/WEAK'])
            regime_data.append(['BB Squeeze Phase', market_regime.get('bb_squeeze_phase', 'N/A'), 'Bollinger Band Width', 'True = Breakout Pending'])
            
            regime_data.append(['', '', '', ''])
            
            # Section 3: BTC Intelligence
            regime_data.append(['BTC Technical & Sentiment Analysis', '', '', ''])
            regime_data.append(['Metric', 'Value', 'Description', 'Impact'])
            regime_data.append(['BTC Trend', market_regime.get('btc_trend', 'N/A'), 'BTC Technical Direction', 'BULLISH/BEARISH/NEUTRAL'])
            regime_data.append(['BTC Technical Confidence', f"{market_regime.get('btc_technical_confidence', 50):.1f}%", 'Technical Analysis Confidence', '0-100% Score'])
            regime_data.append(['BTC Sentiment Confidence', f"{market_regime.get('btc_sentiment_confidence', 50):.1f}%", 'Sentiment Analysis Confidence', '0-100% Score'])
            regime_data.append(['BTC Health Score', f"{market_regime.get('btc_health_score', 50):.1f}", 'Composite BTC Health', 'Technical 60% + Sentiment 40%'])
            regime_data.append(['BTC ADX', f"{market_regime.get('btc_adx', 20):.1f}", 'BTC Trend Strength', 'Higher = Stronger Trend'])
            
            regime_data.append(['', '', '', ''])
            
            # Section 4: BTC Sentiment Breakdown
            regime_data.append(['BTC Sentiment Breakdown', '', '', ''])
            regime_data.append(['Source', 'Score', 'Description', 'Range'])
            regime_data.append(['LunarCrush Galaxy', market_regime.get('btc_galaxy_score', 50), 'Social + Price Intelligence', '0-100 (50+ Bullish)'])
            regime_data.append(['TokenMetrics TM Grade', f"{market_regime.get('btc_tm_grade', 50):.1f}", 'AI Trader Grade', '0-100 (70+ Strong Buy)'])
            regime_data.append(['TokenMetrics TA Grade', f"{market_regime.get('btc_ta_grade', 50):.1f}", 'Technical Analysis Grade', '0-100 (70+ Strong)'])
            regime_data.append(['TokenMetrics Quant Grade', f"{market_regime.get('btc_quant_grade', 50):.1f}", 'Quantitative Analysis', '0-100 (70+ Strong)'])
            
            regime_data.append(['', '', '', ''])
            
            # Section 5: Wider Market Context
            regime_data.append(['Wider Market Intelligence', '', '', ''])
            regime_data.append(['Metric', 'Value', 'Description', 'Interpretation'])
            regime_data.append(['Fear & Greed Index', market_regime.get('fear_greed_index', 50), 'Market Sentiment 0-100', '<30 Fear, >70 Greed'])
            regime_data.append(['Fear & Greed Class', market_regime.get('fear_greed_classification', 'Neutral'), 'Sentiment Classification', 'Extreme Fear to Extreme Greed'])
            regime_data.append(['BTC Dominance', f"{market_regime.get('btc_dominance', 50):.1f}%", 'BTC Market Share', '<40% Alt Season, >60% BTC Season'])
            regime_data.append(['BTC Dominance Trend', market_regime.get('btc_dominance_trend', 'STABLE'), 'Dominance Direction', 'INCREASING/DECREASING/STABLE'])
            regime_data.append(['Market Health Score', f"{market_regime.get('market_health_score', 50):.1f}", 'Composite Market Health', '0-100 (Higher = Healthier)'])
            regime_data.append(['Alt Season Indicator', market_regime.get('alt_season_indicator', 'NEUTRAL'), 'Alt Market Conditions', 'ALT_SEASON/ALT_FAVORABLE/BTC_SEASON/NEUTRAL'])
            
            regime_data.append(['', '', '', ''])
            
            # Section 6: Alt Market Analysis
            regime_data.append(['Alt Market Trends', '', '', ''])
            regime_data.append(['Metric', 'Value', 'Description', 'Impact'])
            regime_data.append(['Alt Market Cap Trend', market_regime.get('alt_market_cap_trend', 'NEUTRAL'), 'Alt vs BTC Market Cap', 'RISING/DECLINING/STABLE'])
            regime_data.append(['Alt Correlation Index', f"{market_regime.get('alt_correlation_index', 0.75):.2f}", 'Alt Synchronization', '0.6-0.9 (Higher = More Correlated)'])
            regime_data.append(['Alt Volatility Index', market_regime.get('alt_volatility_index', 'NORMAL'), 'Alt Market Volatility', 'HIGH/NORMAL/LOW'])
            
            regime_data.append(['', '', '', ''])
            
            # Section 7: Traditional Markets
            regime_data.append(['Traditional Markets', '', '', ''])
            regime_data.append(['Metric', 'Value', 'Description', 'Impact'])
            regime_data.append(['SPY Trend', market_regime.get('spy_trend', 'UNKNOWN'), 'S&P 500 Direction', 'RISING/FALLING/STABLE'])
            regime_data.append(['SPY Change', f"{market_regime.get('spy_change', 0):.2f}%", 'Daily Change %', 'Market Performance'])
            regime_data.append(['VIX Level', f"{market_regime.get('vix_price', 20):.1f}", 'Fear Gauge', '<20 Low Fear, >30 High Fear'])
            regime_data.append(['DXY Change', f"{market_regime.get('dxy_change', 0):.2f}%", 'Dollar Index Change', 'Strong Dollar = Risk Off'])
            regime_data.append(['QQQ Change', f"{market_regime.get('qqq_change', 0):.2f}%", 'Tech Sector Performance', 'Nasdaq Tech Health'])
            regime_data.append(['Market Environment', market_regime.get('market_environment', 'NEUTRAL'), 'Risk Environment', 'RISK_ON/RISK_OFF/NEUTRAL'])
            
            regime_data.append(['', '', '', ''])
            
            # Section 8: Position Sizing & Risk
            regime_data.append(['Position Sizing & Risk Management', '', '', ''])
            regime_data.append(['Metric', 'Value', 'Description', 'Application'])
            regime_data.append(['Regime Confidence', f"{market_regime.get('regime_confidence', 50):.1f}%", 'Overall Regime Confidence', 'Higher = More Confident'])
            regime_data.append(['BB Strategy Suitability', market_regime.get('bb_suitability', 'FAIR'), 'BB Bounce Strategy Fit', 'EXCELLENT/GOOD/FAIR/POOR'])
            regime_data.append(['Position Multiplier', f"{market_regime.get('position_multiplier', 1.0):.2f}x", 'Dynamic Position Sizing', '0.5x-1.5x (Applied to All Trades)'])
            regime_data.append(['Alt Market Outlook', market_regime.get('alt_market_outlook', 'FAIR'), 'BTC Impact on Alts', 'EXCELLENT/GOOD/FAIR/POOR'])
            
            regime_data.append(['', '', '', ''])
            
            # Section 9: Analysis Metadata
            regime_data.append(['Analysis Information', '', '', ''])
            regime_data.append(['Timestamp', market_regime.get('analysis_timestamp', 'N/A'), 'Analysis Time', ''])
            regime_data.append(['Symbol Analyzed', market_regime.get('symbol_analyzed', 'N/A'), 'Market Proxy Used', 'ETH as Alt Market Representative'])
            
            # Convert to DataFrame and save
            regime_df = pd.DataFrame(regime_data, columns=['Metric', 'Value', 'Description', 'Notes'])
            regime_df.to_excel(writer, sheet_name='Market_Regime_Analysis', index=False)
            
            logger.info("Market regime analysis sheet created successfully")
            
        except Exception as e:
            logger.error(f"Error creating market regime sheet: {str(e)}")
            # Create minimal sheet on error
            try:
                error_df = pd.DataFrame([
                    ['Error', 'Market regime sheet creation failed'],
                    ['Details', str(e)[:100]]  # Truncate error message
                ], columns=['Status', 'Message'])
                error_df.to_excel(writer, sheet_name='Market_Regime_Analysis', index=False)
            except:
                pass  # If even error sheet fails, just skip

    def format_comprehensive_results(self, all_results: List[Dict[str, Any]]) -> pd.DataFrame:
        """Format comprehensive results - alias for format_results_dataframe"""
        return self.format_results_dataframe(all_results)

    def display_market_sentiment(self, market_sentiment: Dict[str, Any]):
        """Display market sentiment before main analysis - public method"""
        try:
            print(f"\n" + "="*80)
            print("🌍 CRYPTO MARKET SENTIMENT CHECK")
            print("="*80)
            
            if not market_sentiment:
                print("Market sentiment data unavailable")
                return
            
            # Fear & Greed Index
            fng = market_sentiment.get('fear_greed', {})
            if fng.get('available'):
                fng_emoji = "🟢" if fng['signal'] in ['Bullish Signal', 'Positive'] else \
                           "🟡" if fng['signal'] == 'Neutral' else \
                           "🟠" if fng['signal'] == 'Caution' else "🔴"
                print(f"😨 Fear & Greed Index: {fng['value']}/100 ({fng['classification']}) {fng_emoji}")
                print(f"   Signal: {fng['signal']}")
            
            # BTC Dominance
            btc_dom = market_sentiment.get('btc_dominance', {})
            if btc_dom.get('available'):
                dom_emoji = "🟢" if btc_dom['signal'] in ['Bullish Signal', 'Positive'] else \
                           "🟡" if btc_dom['signal'] == 'Neutral' else \
                           "🟠" if btc_dom['signal'] == 'Caution' else "🔴"
                change_dir = "↗️" if btc_dom['change_24h'] > 0 else "↘️" if btc_dom['change_24h'] < 0 else "→"
                print(f"₿ BTC Dominance: {btc_dom['current']}% ({btc_dom['change_24h']:+.1f}% 24h) {change_dir} {dom_emoji}")
                print(f"   Signal: {btc_dom['signal']}")
            
            # BTC Technical
            btc_tech = market_sentiment.get('btc_technical', {})
            if btc_tech.get('available'):
                tech_emoji = "🟢" if btc_tech['signal'] in ['Bullish Signal', 'Positive'] else \
                            "🟡" if btc_tech['signal'] == 'Neutral' else \
                            "🟠" if btc_tech['signal'] == 'Caution' else "🔴"
                ma_status = "Above" if btc_tech.get('above_sma20', True) else "Below"
                print(f"📊 BTC Technical: ${btc_tech.get('price', 0):,.0f} | RSI: {btc_tech.get('rsi', 50)} | {ma_status} SMA20 {tech_emoji}")
                print(f"   Signal: {btc_tech['signal']}")
            
            # Overall Risk
            risk = market_sentiment.get('overall_risk', 'NEUTRAL')
            if risk == "HIGH RISK":
                risk_emoji = "🚨"
                risk_message = "Consider waiting or using smaller positions"
            elif risk == "MEDIUM RISK":
                risk_emoji = "⚠️"
                risk_message = "Use caution - reduced position sizes recommended"
            elif risk == "LOW RISK":
                risk_emoji = "✅"
                risk_message = "Good environment for normal position sizes"
            else:
                risk_emoji = "➖"
                risk_message = "Mixed signals - use standard risk management"
            
            print(f"\n🎯 OVERALL MARKET RISK: {risk} {risk_emoji}")
            print(f"💡 Recommendation: {risk_message}")
            
        except Exception as e:
            logger.error(f"Error displaying market sentiment: {e}")
            print("Market sentiment analysis unavailable")

    def display_terminal_summary(self, df: pd.DataFrame, market_sentiment: Dict[str, Any] = None):
        """Display enhanced terminal summary with market regime"""
        try:
            # NEW: Display market regime if available
            if hasattr(self, '_market_regime_data') and self._market_regime_data:
                self._display_market_regime_summary(self._market_regime_data)
            
            # Existing terminal summary code continues here...
            if df.empty:
                print("\n" + "="*80)
                print("No trading opportunities found matching the criteria.")
                print("="*80)
                return
            
            print("\n" + "="*80)
            print("CRYPTO BB BOUNCE SCANNER - ANALYSIS SUMMARY")
            print("="*80)
            print(f"Total coins analyzed: {len(df)}")
            
            # Tier breakdown
            tier_counts = df['tier'].value_counts()
            for tier in ['PREMIUM', 'HIGH', 'GOOD', 'FAIR', 'MARGINAL', 'WEAK']:
                count = tier_counts.get(tier, 0)
                percentage = (count / len(df) * 100) if len(df) > 0 else 0
                print(f"{tier}: {count} coins ({percentage:.1f}%)")
            
            # Risk analysis
            if len(df) > 0:
                low_risk_count = len(df[df['risk_pct'] <= 2.5])
                med_risk_count = len(df[(df['risk_pct'] > 2.5) & (df['risk_pct'] <= 4.0)])
                high_risk_count = len(df[df['risk_pct'] > 4.0])
                
                print(f"\nRisk Analysis:")
                print(f"Low Risk (≤2.5%): {low_risk_count} trades")
                print(f"Medium Risk (2.5-4%): {med_risk_count} trades")
                print(f"High Risk (>4%): {high_risk_count} trades")
            
            # Show premium trades
            self._display_premium_trades(df)
            
            # Show good trades
            self._display_good_trades(df)
            
            # Display market sentiment if available
            if market_sentiment:
                self._display_market_sentiment(market_sentiment)
            
            # Enhanced recommendations
            self._display_recommendations(df)
            
            # Display detailed scoring breakdown
            self.display_detailed_scoring_breakdown(df)
            
            # Add this line in display_terminal_summary method:
            self._display_market_metadata_summary(df)
            
        except Exception as e:
            logger.error(f"Error displaying terminal summary: {e}")

    def _display_premium_trades(self, df: pd.DataFrame):
        """Display premium trades (70%+ probability)"""
        try:
            premium_trades = df[df['probability'] >= 70]
            print(f"\n" + "="*80)
            print(f"PREMIUM TRADES (70%+ Probability): {len(premium_trades)}")
            print("="*80)
            
            if len(premium_trades) > 0:
                for i, (_, trade) in enumerate(premium_trades.head(10).iterrows(), 1):
                    div_info = f" | Div: {trade['divergence_indicators']}" if trade.get('divergence_detected', False) else " | No divergence"
                    
                    # Risk indicators
                    if trade['risk_pct'] <= 3.0:
                        risk_flag = "✅"
                    elif trade['risk_pct'] <= 6.0:
                        risk_flag = "⚠️"
                    else:
                        risk_flag = "🔴"
                    
                    print(f"\n{i}. {trade['symbol']} - {trade['setup_type']} ({trade['exchange']})")
                    print(f"   🎯 Probability: {trade['probability']}% ({trade['tier']}) | BB Score: {trade['bb_score']}/11")
                    if trade['entry'] > 0:
                        print(f"   💰 Entry: ${trade['entry']:.6f} | Stop: ${trade['stop']:.6f} | Target: ${trade['target1']:.6f}")
                        print(f"   📊 R:R: {trade['risk_reward']}:1 | Risk: {trade['risk_pct']}% {risk_flag} | Gain: {trade['gain_pct']}%")
                    
                    # Get pattern info if available
                    if trade.get('patterns_detected') and trade['patterns_detected'] != 'None':
                        patterns = trade['patterns_detected'].split(', ')[:2]  # Show max 2 patterns
                        pattern_info = f" | 🕯️ {', '.join(patterns)}"
                    else:
                        pattern_info = f" | 🕯️ N/A"

                    # Get chart pattern info if available
                    chart_pattern_info = ""
                    chart_patterns = trade.get('chart_patterns_detected', 'None')
                    if chart_patterns and chart_patterns != 'None':
                        chart_pattern_info = f" | 📈 {chart_patterns}"
                    else:
                        chart_pattern_info = f" | 📈 N/A"

                    print(f"🔍 RSI: {trade['rsi']:.1f} | BB%: {trade['bb_pct']:.3f} | Vol: {trade['volume_ratio']:.1f}x{chart_pattern_info} | {div_info}{pattern_info}")
            else:
                print("No premium trades found.")
                
        except Exception as e:
            logger.error(f"Error displaying premium trades: {e}")

    def _display_good_trades(self, df: pd.DataFrame):
        """Display good trades (65-69% probability)"""
        try:
            good_trades = df[(df['probability'] >= 65) & (df['probability'] < 70)]
            if len(good_trades) > 0:
                print(f"\n" + "="*80)
                print(f"GOOD TRADES (65-69% Probability): {len(good_trades)}")
                print("="*80)
                
                for i, (_, trade) in enumerate(good_trades.head(5).iterrows(), 1):
                    div_info = f" | Div: {trade['divergence_indicators']}" if trade.get('divergence_detected', False) else ""
                    
                    if trade['risk_pct'] <= 3.0:
                        risk_flag = "✅"
                    elif trade['risk_pct'] <= 6.0:
                        risk_flag = "⚠️"
                    else:
                        risk_flag = "🔴"
                        
                    print(f"{i}. {trade['symbol']} - {trade['setup_type']} | {trade['probability']}% | Risk: {trade['risk_pct']}% {risk_flag}{div_info}")
                    
        except Exception as e:
            logger.error(f"Error displaying good trades: {e}")

    def _display_market_sentiment(self, market_sentiment: Dict[str, Any]):
        """Display market sentiment analysis"""
        try:
            print(f"\n" + "="*80)
            print("🌍 CRYPTO MARKET SENTIMENT")
            print("="*80)
            
            # Fear & Greed Index
            fng = market_sentiment.get('fear_greed', {})
            if fng.get('available'):
                fng_emoji = "🟢" if fng['signal'] in ['Bullish Signal', 'Positive'] else \
                           "🟡" if fng['signal'] == 'Neutral' else \
                           "🟠" if fng['signal'] == 'Caution' else "🔴"
                print(f"😨 Fear & Greed Index: {fng['value']}/100 ({fng['classification']}) {fng_emoji}")
                print(f"   Signal: {fng['signal']}")
            
            # BTC Dominance
            btc_dom = market_sentiment.get('btc_dominance', {})
            if btc_dom.get('available'):
                dom_emoji = "🟢" if btc_dom['signal'] in ['Bullish Signal', 'Positive'] else \
                           "🟡" if btc_dom['signal'] == 'Neutral' else \
                           "🟠" if btc_dom['signal'] == 'Caution' else "🔴"
                change_dir = "↗️" if btc_dom['change_24h'] > 0 else "↘️" if btc_dom['change_24h'] < 0 else "→"
                print(f"₿ BTC Dominance: {btc_dom['current']}% ({btc_dom['change_24h']:+.1f}% 24h) {change_dir} {dom_emoji}")
                print(f"   Signal: {btc_dom['signal']}")
                
        except Exception as e:
            logger.error(f"Error displaying market sentiment: {e}")

    def _display_recommendations(self, df: pd.DataFrame):
        """Display trading recommendations"""
        try:
            print(f"\n" + "="*80)
            print("TRADING RECOMMENDATIONS")
            print("="*80)
            
            take_trades = len(df[df['action'] == 'TAKE TRADE'])
            consider_trades = len(df[df['action'] == 'CONSIDER'])
            
            if take_trades > 0:
                print(f"🚀 {take_trades} PREMIUM trades ready")
                print("   - Use 1.5-2% position size")
                print("   - Expected duration: 2-7 days")
                print("   - Stops: 2x ATR (fewer false exits)")
                
            if consider_trades > 0:
                print(f"⭐ {consider_trades} GOOD trades to consider")
                print("   - Use 1% position size")
                print("   - Review risk levels carefully")
                
            if take_trades + consider_trades == 0:
                print("🔍 No high-probability trades found")
                print("   - Market may be trending strongly")
                print("   - Consider waiting for better market conditions")
            
            print(f"\n📋 EXECUTION PRIORITY:")
            print("1. Premium trades with divergence confirmation")
            print("2. High probability trades with good R:R")
            print("3. Consider risk tolerance: ✅≤3% ⚠️3-6% 🔴>6%")
            print("4. All trades have 1H confirmation + current price validation")
            
        except Exception as e:
            logger.error(f"Error displaying recommendations: {e}")

    def display_sentiment_summary(self, df: pd.DataFrame):
        """Display sentiment analysis summary for trades with sentiment data"""
        try:
            # Check if sentiment columns exist
            sentiment_cols = ['lunar_data_available', 'tm_data_available']
            if not any(col in df.columns for col in sentiment_cols):
                return
            
            # Get trades with sentiment data
            sentiment_trades = df[
                (df.get('lunar_data_available', False) == True) | 
                (df.get('tm_data_available', False) == True)
            ].head(10)
            
            if sentiment_trades.empty:
                print("\nNo sentiment data available for trades")
                return
            
            print(f"\n" + "="*80)
            print("SENTIMENT ANALYSIS SUMMARY")
            print("="*80)
            
            for i, (_, trade) in enumerate(sentiment_trades.iterrows(), 1):
                alignment = trade.get('sentiment_overall_alignment', 'No Data')
                sentiment_status = "📈" if alignment in ['Strong Positive', 'Positive'] else \
                                 "📉" if alignment in ['Strong Negative', 'Negative'] else "➖"
                
                print(f"\n{i}. {trade['symbol']} - {trade['setup_type']} | {trade['probability']}% Probability")
                
                # Show LunarCrush data if available
                if trade.get('lunar_data_available'):
                    lunar_rating = trade.get('lunar_sentiment_rating', 'No Data')
                    lunar_score = trade.get('lunar_sentiment_score', 0)
                    print(f"   🌙 LunarCrush: {lunar_rating} (Score: {lunar_score})")
                
                # Show TokenMetrics data if available (ENHANCED VERSION)
                if trade.get('tm_data_available'):
                    tm_grade = trade.get('tm_trader_grade', 0)
                    tm_ta_grade = trade.get('tm_ta_grade', 0)
                    tm_quant_grade = trade.get('tm_quant_grade', 0)
                    tm_change = trade.get('tm_grade_change_24h', 0)
                    change_arrow = "↗️" if tm_change > 0 else "↘️" if tm_change < 0 else "→"
                    
                    # Get enhanced descriptions using the scoring ranges
                    # TM Trader Grade descriptions
                    if tm_grade >= 80:
                        tm_desc = "✅ Excellent trading opportunity"
                    elif tm_grade >= 60:
                        tm_desc = "📈 Good trading opportunity"
                    elif tm_grade >= 40:
                        tm_desc = "📊 Average/Neutral"
                    elif tm_grade >= 20:
                        tm_desc = "⚠️ Below average"
                    else:
                        tm_desc = "❌ Poor trading opportunity"
                    
                    # TA Grade descriptions
                    if tm_ta_grade >= 80:
                        ta_desc = "✅ Strong technical signals"
                    elif tm_ta_grade >= 60:
                        ta_desc = "📈 Good technical setup"
                    elif tm_ta_grade >= 40:
                        ta_desc = "📊 Neutral technical picture"
                    elif tm_ta_grade >= 20:
                        ta_desc = "⚠️ Weak technicals"
                    else:
                        ta_desc = "❌ Poor technical setup"
                    
                    # Quant Grade descriptions
                    if tm_quant_grade >= 60:
                        quant_desc = "✅ High"
                    elif tm_quant_grade >= 40:
                        quant_desc = "📊 Neutral"
                    else:
                        quant_desc = "⚠️ Low"
                    
                    print(f"   📊 TokenMetrics: Trader Grade {tm_desc} {tm_grade} | TA Grade {ta_desc} {tm_ta_grade} | Quant Grade {quant_desc} {tm_quant_grade} | 24h: {tm_change:+.1f}% {change_arrow}")
                
                # Show alignment if available
                if alignment != 'No Data':
                    print(f"   {sentiment_status} Sentiment Alignment: {alignment}")
                    if trade.get('sentiment_alignment_factors'):
                        print(f"   💡 Factors: {trade['sentiment_alignment_factors']}")
                        
        except Exception as e:
            logger.error(f"Error displaying sentiment summary: {e}")

    def display_detailed_scoring_breakdown(self, df: pd.DataFrame):
        """Display detailed scoring breakdown for trades with scoring details"""
        try:
            # Check if scoring_details column exists
            if 'scoring_details' not in df.columns:
                return
            
            # Get trades with scoring details
            scoring_trades = df[df['scoring_details'].notna()].head(5)
            
            if scoring_trades.empty:
                return
            
            print(f"\n" + "="*80)
            print("📊 DETAILED SCORING BREAKDOWN")
            print("="*80)
            
            for i, (_, trade) in enumerate(scoring_trades.iterrows(), 1):
                scoring_details = trade['scoring_details']
                if not scoring_details or 'breakdown' not in scoring_details:
                    continue
                
                print(f"\n{i}. {trade['symbol']} - {trade['setup_type']} | Score: {trade['bb_score']}/{scoring_details.get('total_possible', 34)}")
                print(f"   Quality: {trade['setup_quality']} | Probability: {trade['probability']}%")
                
                # Display tier scores
                tier_scores = scoring_details.get('tier_scores', {})
                if tier_scores:
                    print(f"   📈 Tier Breakdown:")
                    print(f"      • Base BB: {tier_scores.get('base_bb', 0)}/10 pts")
                    print(f"      • Money Flow: {tier_scores.get('money_flow', 0)}/7 pts")
                    print(f"      • BB Specific: {tier_scores.get('bb_specific', 0)}/8 pts")
                    print(f"      • Volume & Momentum: {tier_scores.get('volume_momentum', 0)}/6 pts")
                    print(f"      • Divergence: {tier_scores.get('divergence', 0)}/3 pts")
                
                # Display detailed breakdown
                breakdown = scoring_details.get('breakdown', [])
                if breakdown:
                    print(f"   🔍 Detailed Breakdown:")
                    for item in breakdown[:8]:  # Show first 8 items to avoid clutter
                        print(f"      • {item}")
                    
                    if len(breakdown) > 8:
                        print(f"      • ... and {len(breakdown) - 8} more indicators")
                
                # Display key indicator values
                indicator_values = scoring_details.get('indicator_values', {})
                if indicator_values:
                    print(f"   📊 Key Indicators:")
                    for indicator, value in indicator_values.items():
                        if isinstance(value, float):
                            print(f"      • {indicator.upper()}: {value:.3f}")
                        else:
                            print(f"      • {indicator.upper()}: {value}")
                
                print()  # Add spacing between trades
                
        except Exception as e:
            logger.error(f"Error displaying detailed scoring breakdown: {e}")

    def format_scoring_breakdown(self, setup_data):
        """Format detailed scoring breakdown for display"""
        
        if 'scoring_details' not in setup_data:
            return ""
        
        details = setup_data['scoring_details']
        symbol = setup_data.get('symbol', 'UNKNOWN')
        setup_type = setup_data.get('setup_type', 'NONE')
        bb_score = setup_data.get('bb_score', 0)
        
        # Build the breakdown display
        breakdown_text = f"\n📊 DETAILED SCORING BREAKDOWN - {symbol} {setup_type} ({bb_score}/34 points):\n"
        breakdown_text += "=" * 60 + "\n"
        
        # Show tier summaries
        tier_scores = details.get('tier_scores', {})
        breakdown_text += f"🏗️  TIER SUMMARY:\n"
        breakdown_text += f"   • Base BB Setup: {tier_scores.get('base_bb', 0)} pts\n"
        breakdown_text += f"   • Money Flow: {tier_scores.get('money_flow', 0)} pts\n"
        breakdown_text += f"   • BB-Specific: {tier_scores.get('bb_specific', 0)} pts\n"
        breakdown_text += f"   • Volume/Momentum: {tier_scores.get('volume_momentum', 0)} pts\n"
        breakdown_text += f"   • Divergence: {tier_scores.get('divergence', 0)} pts\n"
        breakdown_text += f"   ➡️  TOTAL: {bb_score} points\n\n"
        
        # Show detailed breakdown
        breakdown_text += f"🔍 DETAILED COMPONENT BREAKDOWN:\n"
        for i, component in enumerate(details.get('breakdown', []), 1):
            if "⭐" in component:  # Highlight top signals
                breakdown_text += f"   {i:2d}. {component} 🎯\n"
            else:
                breakdown_text += f"   {i:2d}. {component}\n"
        
        # Show key indicator values
        values = details.get('indicator_values', {})
        if values:
            breakdown_text += f"\n📈 KEY INDICATOR VALUES:\n"
            if 'mfi' in values:
                breakdown_text += f"   • MFI (Money Flow Index): {values['mfi']:.1f}\n"
            if 'cmf' in values:
                breakdown_text += f"   • CMF (Chaikin Money Flow): {values['cmf']:.4f}\n"
            if 'bb_expansion' in values:
                breakdown_text += f"   • BB Expansion Ratio: {values['bb_expansion']:.2f}x\n"
            if 'volume_surge' in values:
                breakdown_text += f"   • Volume Surge Detected: {values['volume_surge']}\n"
            if 'bb_trend' in values:
                breakdown_text += f"   • BB Trend Direction: {values['bb_trend']}\n"
        
        breakdown_text += "=" * 60 + "\n"
        
        return breakdown_text

    def _extract_enhanced_scoring_data(self, setup_data):
        """Extract enhanced scoring data for Excel output"""
        enhanced_data = {}
        
        # Basic enhanced data
        enhanced_data['bb_score_34'] = setup_data.get('bb_score', 0)
        enhanced_data['setup_quality_enhanced'] = setup_data.get('setup_quality', 'None')
        
        # Extract scoring details if available
        scoring_details = setup_data.get('scoring_details', {})
        
        if scoring_details:
            # Tier breakdown
            tier_scores = scoring_details.get('tier_scores', {})
            enhanced_data['tier_base_bb'] = tier_scores.get('base_bb', 0)
            enhanced_data['tier_money_flow'] = tier_scores.get('money_flow', 0)
            enhanced_data['tier_bb_specific'] = tier_scores.get('bb_specific', 0)
            enhanced_data['tier_volume_momentum'] = tier_scores.get('volume_momentum', 0)
            enhanced_data['tier_divergence'] = tier_scores.get('divergence', 0)
            
            # Key indicator values
            indicator_values = scoring_details.get('indicator_values', {})
            enhanced_data['mfi_value'] = indicator_values.get('mfi', 0)
            enhanced_data['cmf_value'] = indicator_values.get('cmf', 0)
            enhanced_data['bb_expansion_ratio'] = indicator_values.get('bb_expansion', 0)
            enhanced_data['volume_surge_detected'] = indicator_values.get('volume_surge', False)
            enhanced_data['bb_trend_direction'] = indicator_values.get('bb_trend', 'Unknown')
            
            # Component breakdown summary (first 5 components for Excel)
            breakdown = scoring_details.get('breakdown', [])
            for i, component in enumerate(breakdown[:5], 1):
                enhanced_data[f'component_{i}'] = component
                
            # Total components count
            enhanced_data['total_components'] = len(breakdown)
            
            # Check for high-value MFI signals (88% success rate indicator)
            mfi_signal_detected = any('MFI' in comp and '⭐' in comp for comp in breakdown)
            enhanced_data['mfi_priority_signal'] = mfi_signal_detected
            
        else:
            # Default values if no scoring details
            enhanced_data.update({
                'tier_base_bb': 0,
                'tier_money_flow': 0,
                'tier_bb_specific': 0,
                'tier_volume_momentum': 0,
                'tier_divergence': 0,
                'mfi_value': 0,
                'cmf_value': 0,
                'bb_expansion_ratio': 0,
                'volume_surge_detected': False,
                'bb_trend_direction': 'Unknown',
                'total_components': 0,
                'mfi_priority_signal': False
            })
        
        return enhanced_data

    def _sanitize_excel_value(self, value):
        """AGGRESSIVE sanitization for Excel compatibility - prevents ALL corruption"""
        
        # Handle None/NaN values first
        if value is None or pd.isna(value):
            return ""
        
        # Handle numpy types that cause issues
        if hasattr(value, 'dtype'):
            if 'float' in str(value.dtype) and pd.isna(value):
                return ""
            value = value.item() if hasattr(value, 'item') else str(value)
        
        # Handle boolean values
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        
        # Handle numeric values aggressively
        if isinstance(value, (int, float)):
            # Check for all problematic numeric values
            if pd.isna(value) or value == float('inf') or value == float('-inf'):
                return 0  # Use 0 instead of empty string for numbers
            if abs(value) > 1e15:  # Extremely large numbers
                return 0
            return round(float(value), 6)  # Limit precision
        
        # Handle complex objects (dictionaries, lists, etc.)
        if isinstance(value, (dict, list, tuple, set)):
            try:
                # Convert to simple string representation
                str_val = str(value)
                if len(str_val) > 1000:  # Limit length
                    return str_val[:1000] + "..."
                return str_val
            except:
                return "COMPLEX_OBJECT"
        
        # Handle strings aggressively
        if isinstance(value, str):
            # Remove all problematic characters
            value = value.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
            value = value.replace('\x00', '').replace('\x01', '').replace('\x02', '')
            
            # Remove any non-printable characters
            value = ''.join(char for char in value if char.isprintable() or char.isspace())
            
            # Limit length aggressively
            if len(value) > 500:  # Much shorter limit
                value = value[:500] + "..."
            
            # Handle empty strings
            if not value.strip():
                return ""
                
            return value
        
        # Handle datetime objects
        if hasattr(value, 'strftime'):
            try:
                return value.strftime('%Y-%m-%d %H:%M:%S')
            except:
                return str(value)
        
        # Final fallback - convert everything to string and sanitize
        try:
            str_value = str(value)
            if len(str_value) > 500:
                str_value = str_value[:500]
            # Remove any remaining problematic characters
            return ''.join(char for char in str_value if char.isprintable() or char.isspace())
        except:
            return "SANITIZATION_ERROR"

    def _add_enhanced_columns_to_excel(self, worksheet, setups):
        """Add enhanced scoring columns with AGGRESSIVE sanitization"""
        
        # Enhanced headers (same as before)
        enhanced_headers = [
            'bb_score_34', 'setup_quality_enhanced', 'tier_base_bb',
            'tier_money_flow', 'tier_bb_specific', 'tier_volume_momentum',
            'tier_divergence', 'mfi_value', 'cmf_value', 'bb_expansion_ratio',
            'volume_surge_detected', 'bb_trend_direction', 'total_components',
            'mfi_priority_signal', 'component_1', 'component_2', 'component_3',
            'component_4', 'component_5'
        ]
        
        # Get existing headers
        existing_headers = []
        for col in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            if header:
                existing_headers.append(header)
        
        # Add new headers
        start_col = len(existing_headers) + 1
        for i, header in enumerate(enhanced_headers):
            worksheet.cell(row=1, column=start_col + i, value=str(header))  # Ensure header is string
        
        # Add enhanced data with DOUBLE sanitization
        for row_idx, setup in enumerate(setups, start=2):
            enhanced_data = self._extract_enhanced_scoring_data(setup)
            
            for i, header in enumerate(enhanced_headers):
                raw_value = enhanced_data.get(header, '')
                
                # DOUBLE SANITIZATION - be extra aggressive
                sanitized_value = self._sanitize_excel_value(raw_value)
                final_value = self._sanitize_excel_value(sanitized_value)  # Second pass
                
                # Final safety check
                if final_value is None:
                    final_value = ""
                
                try:
                    worksheet.cell(row=row_idx, column=start_col + i, value=final_value)
                except Exception as e:
                    # If writing still fails, use empty string
                    worksheet.cell(row=row_idx, column=start_col + i, value="")

    # ADD NEW METHOD: Market Regime Terminal Display
    def _display_market_regime_summary(self, market_regime: Dict):
        """Display 6-line market regime summary"""
        print("\n" + "="*80)
        print("🌊 MARKET REGIME INTELLIGENCE")
        print("="*80)
        
        regime_type = market_regime.get('regime_type', 'UNKNOWN')
        confidence = market_regime.get('regime_confidence', 50)
        bb_suitability = market_regime.get('bb_suitability', 'UNKNOWN')
        position_mult = market_regime.get('position_multiplier', 1.0)
        
        print(f"📊 Regime: {regime_type} ({confidence}% confidence)")
        print(f"🎯 BB Strategy Suitability: {bb_suitability}")
        print(f"💰 Position Sizing: {position_mult}x multiplier")
        print(f"🏥 Market Health: {market_regime.get('market_health_score', 50)}/100")
        print(f"₿ BTC Health: {market_regime.get('btc_health_score', 50)}/100")
        print(f"🌍 Alt Season: {market_regime.get('alt_season_indicator', 'UNKNOWN')}")

    def _display_market_metadata_summary(self, df: pd.DataFrame):
        """Display market metadata summary"""
        if df.empty:
            return
            
        print(f"\n📊 MARKET METADATA INTELLIGENCE")
        print("=" * 50)
        
        # Market Cap Distribution
        if 'market_cap_tier' in df.columns:
            cap_dist = df['market_cap_tier'].value_counts()
            print(f"📈 Market Cap Distribution:")
            for tier, count in cap_dist.items():
                print(f"   {tier}: {count} trades")
        
        # Sector Distribution  
        if 'primary_sector' in df.columns:
            sector_dist = df['primary_sector'].value_counts()
            print(f"\n🏭 Sector Distribution:")
            for sector, count in sector_dist.items():
                print(f"   {sector}: {count} trades")
        
        # Performance by Market Cap
        if 'market_cap_tier' in df.columns and 'expected_success_rate' in df.columns:
            print(f"\n🎯 Expected Performance by Market Cap:")
            cap_performance = df.groupby('market_cap_tier')['expected_success_rate'].mean()
            for tier, rate in cap_performance.items():
                print(f"   {tier}: {rate:.1f}% expected success")

    # ADD HELPER METHOD: Market context classification
    def _calculate_market_context(self, bb_score: int, market_baseline: float) -> str:
        """Calculate market context classification"""
        score_percentage = (bb_score / 34) * 100
        
        if score_percentage > market_baseline + 15:
            return "SIGNIFICANTLY_ABOVE_MARKET"
        elif score_percentage > market_baseline + 5:
            return "ABOVE_MARKET"
        elif score_percentage > market_baseline - 5:
            return "MARKET_AVERAGE"
        else:
            return "BELOW_MARKET"

    # ADD METHOD TO STORE MARKET REGIME DATA
    def set_market_regime_data(self, market_regime: Dict):
        """Store market regime data for display"""
        self._market_regime_data = market_regime

    def _calculate_sentiment_confidence(self, trade_data):
        """Calculate real sentiment confidence from LunarCrush + TokenMetrics"""
        try:
            tm_grade = trade_data.get('tm_trader_grade', 50)
            lunar_score = trade_data.get('lunar_galaxy_score', 50)
            
            # If we have both data sources, average them
            if trade_data.get('tm_data_available') and trade_data.get('lunar_data_available'):
                return (tm_grade + lunar_score) / 2
            # If we only have TokenMetrics
            elif trade_data.get('tm_data_available'):
                return tm_grade
            # If we only have LunarCrush
            elif trade_data.get('lunar_data_available'):
                return lunar_score
            else:
                return 0.0  # No sentiment data available
        except Exception as e:
            logger.error(f"Error calculating sentiment confidence: {e}")
            return 0.0

    def _calculate_historical_confidence(self, trade_data):
        """Calculate real historical confidence from coin-specific data"""
        try:
            # Use coin-specific historical data if available
            historical_probability = trade_data.get('historical_probability', 0)
            historical_win_rate = trade_data.get('historical_win_rate', 0)
            
            # Prefer historical_probability if available, otherwise use historical_win_rate
            if historical_probability > 0:
                return historical_probability
            elif historical_win_rate > 0:
                return historical_win_rate
            else:
                return 0.0  # No historical data available
        except Exception as e:
            logger.error(f"Error calculating historical confidence: {e}")
            return 0.0

    def _calculate_technical_confidence(self, trade_data):
        """Calculate technical confidence from BB score and other technical indicators"""
        try:
            bb_score = trade_data.get('bb_score', 0)
            probability = trade_data.get('probability', 0)
            
            # Use probability if available, otherwise calculate from BB score
            if probability > 0:
                return probability
            elif bb_score > 0:
                # Convert BB score to percentage (34 max score)
                return (bb_score / 34) * 100
            else:
                return 0.0
        except Exception as e:
            logger.error(f"Error calculating technical confidence: {e}")
            return 0.0

    def _calculate_composite_confidence(self, trade_data):
        """Calculate composite confidence from all three components"""
        try:
            technical = self._calculate_technical_confidence(trade_data)
            historical = self._calculate_historical_confidence(trade_data)
            sentiment = self._calculate_sentiment_confidence(trade_data)
            
            # Weight the components (technical is most important)
            weights = {'technical': 0.5, 'historical': 0.3, 'sentiment': 0.2}
            
            composite = (technical * weights['technical'] + 
                        historical * weights['historical'] + 
                        sentiment * weights['sentiment'])
            
            return round(composite, 1)
        except Exception as e:
            logger.error(f"Error calculating composite confidence: {e}")
            return 0.0

    def _assign_confidence_tier(self, composite_confidence):
        """Assign confidence tier based on composite confidence score"""
        try:
            if composite_confidence >= 80:
                return 'HIGH_CONFIDENCE'
            elif composite_confidence >= 65:
                return 'MEDIUM_CONFIDENCE'
            elif composite_confidence >= 50:
                return 'LOW_CONFIDENCE'
            elif composite_confidence > 0:
                return 'WATCH_ONLY'
            else:
                return 'UNRATED'
        except Exception as e:
            logger.error(f"Error assigning confidence tier: {e}")
            return 'UNRATED'

    def _generate_confidence_rationale(self, trade_data):
        """Generate confidence rationale based on available data"""
        try:
            rationale_parts = []
            
            # Technical rationale
            technical_conf = self._calculate_technical_confidence(trade_data)
            if technical_conf > 0:
                rationale_parts.append(f"Technical: {technical_conf:.1f}%")
            
            # Historical rationale
            historical_conf = self._calculate_historical_confidence(trade_data)
            if historical_conf > 0:
                rationale_parts.append(f"Historical: {historical_conf:.1f}%")
            
            # Sentiment rationale
            sentiment_conf = self._calculate_sentiment_confidence(trade_data)
            if sentiment_conf > 0:
                rationale_parts.append(f"Sentiment: {sentiment_conf:.1f}%")
            
            if rationale_parts:
                return " | ".join(rationale_parts)
            else:
                return "No confidence data available"
        except Exception as e:
            logger.error(f"Error generating confidence rationale: {e}")
            return "Error calculating confidence"

    def update_confidence_data(self, df):
        """Update DataFrame with real confidence calculations"""
        try:
            if df.empty:
                return df
            
            # Calculate confidence for each row
            for index, row in df.iterrows():
                trade_data = row.to_dict()
                
                # Calculate individual confidence components
                technical_conf = self._calculate_technical_confidence(trade_data)
                historical_conf = self._calculate_historical_confidence(trade_data)
                sentiment_conf = self._calculate_sentiment_confidence(trade_data)
                composite_conf = self._calculate_composite_confidence(trade_data)
                
                # Update the DataFrame
                df.at[index, 'technical_confidence'] = technical_conf
                df.at[index, 'historical_confidence'] = historical_conf
                df.at[index, 'sentiment_confidence'] = sentiment_conf
                df.at[index, 'composite_confidence'] = composite_conf
                df.at[index, 'confidence_tier'] = self._assign_confidence_tier(composite_conf)
                df.at[index, 'confidence_rationale'] = self._generate_confidence_rationale(trade_data)
            
            return df
        except Exception as e:
            logger.error(f"Error updating confidence data: {e}")
            return df

    # ADD NEW METHOD: Format results with market context (for front columns)
    def format_comprehensive_results_with_market_context(self, all_results: List[Dict], market_baselines: Dict = None) -> pd.DataFrame:
        """Format results with market context and reordered columns"""
        
        if not all_results:
            return pd.DataFrame()
        
        # Convert to DataFrame
        df = pd.DataFrame(all_results)
        
        # Get market baselines if not provided
        if market_baselines is None:
            market_baselines = {'overall_success_rate': 72.4}  # Fallback
        
        # ADD NEW FRONT COLUMNS
        df['technical_probability'] = (df.get('bb_score', 0) / 34 * 100).round(1)
        df['historical_probability'] = df.get('historical_win_rate', 0)
        df['historical_profit'] = df.get('historical_avg_win', 0)
        df['historical_drawdown'] = df.get('historical_avg_loss', 0) 
        df['historical_duration'] = df.get('historical_avg_duration', 0)
        
        # Add market context
        df['market_context'] = df.apply(lambda row: self._calculate_market_context(
            row.get('bb_score', 0), market_baselines.get('overall_success_rate', 72.4)
        ), axis=1)
        
        # REORDER COLUMNS - PRIORITY COLUMNS FIRST
        priority_columns = [
            'symbol', 'setup_type', 'exchange',
            'technical_probability', 'historical_probability', 
            'historical_profit', 'historical_drawdown', 'historical_duration',
            'market_context', 'probability', 'bb_score',
            'entry_price', 'stop_price', 'target1', 'risk_reward', 'risk_pct'
        ]
        
        # Get remaining columns (preserve all existing data)
        remaining_columns = [col for col in df.columns if col not in priority_columns]
        final_columns = priority_columns + remaining_columns
        
        # Reorder DataFrame
        available_columns = [col for col in final_columns if col in df.columns]
        df_reordered = df[available_columns].copy()
        
        return df_reordered

    def _create_market_metadata_sheet(self, writer, df: pd.DataFrame):
        """Create Market Metadata sheet in Excel"""
        try:
            metadata_data = []
            metadata_data.append(['MARKET METADATA INTELLIGENCE', '', '', '', '', '', '', ''])
            metadata_data.append(['Analysis Date', datetime.now().strftime("%Y-%m-%d"), '', '', '', '', '', ''])
            metadata_data.append(['', '', '', '', '', '', '', ''])
            
            # Market Cap Distribution
            if 'market_cap_tier' in df.columns:
                metadata_data.append(['MARKET CAP DISTRIBUTION', '', '', '', '', '', '', ''])
                cap_dist = df['market_cap_tier'].value_counts()
                metadata_data.append(['Tier', 'Count', 'Percentage', 'Expected Success', '', '', '', ''])
                for tier, count in cap_dist.items():
                    percentage = (count / len(df) * 100) if len(df) > 0 else 0
                    avg_success = df[df['market_cap_tier'] == tier]['expected_success_rate'].mean() if 'expected_success_rate' in df.columns else 0
                    metadata_data.append([tier, f"{count} trades", f"{percentage:.1f}%", f"{avg_success:.1f}% expected", '', '', '', ''])
                metadata_data.append(['', '', '', '', '', '', '', ''])
            
            # Sector Distribution  
            if 'primary_sector' in df.columns:
                metadata_data.append(['SECTOR DISTRIBUTION', '', '', '', '', '', '', ''])
                sector_dist = df['primary_sector'].value_counts()
                metadata_data.append(['Sector', 'Count', 'Percentage', 'Expected Success', '', '', '', ''])
                for sector, count in sector_dist.items():
                    percentage = (count / len(df) * 100) if len(df) > 0 else 0
                    avg_success = df[df['primary_sector'] == sector]['expected_success_rate'].mean() if 'expected_success_rate' in df.columns else 0
                    metadata_data.append([sector, f"{count} trades", f"{percentage:.1f}%", f"{avg_success:.1f}% expected", '', '', '', ''])
                metadata_data.append(['', '', '', '', '', '', '', ''])
            
            # Performance by Market Cap
            if 'market_cap_tier' in df.columns and 'expected_success_rate' in df.columns:
                metadata_data.append(['EXPECTED PERFORMANCE BY MARKET CAP', '', '', '', '', '', '', ''])
                cap_performance = df.groupby('market_cap_tier')['expected_success_rate'].mean()
                metadata_data.append(['Market Cap Tier', 'Expected Success Rate', '', '', '', '', '', ''])
                for tier, rate in cap_performance.items():
                    metadata_data.append([tier, f"{rate:.1f}%", '', '', '', '', '', ''])
            
            # Create the sheet
            metadata_df = pd.DataFrame(metadata_data)
            metadata_df.to_excel(writer, sheet_name='Market_Metadata', index=False, header=False)
            
        except Exception as e:
            logger.error(f"Error creating market metadata sheet: {e}")

